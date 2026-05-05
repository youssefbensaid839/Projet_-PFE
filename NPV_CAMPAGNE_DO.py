"""
Script d'import Excel → Base de données
========================================
Colonnes Excel : NOM_CAMPAGNE | NPV | DO
Tables cibles   : Campagne.nom | NPV.numero | DonneurdOrdre.nom
Relations       : Campagne.do_id → DonneurdOrdre.id
                  NPV.campagne_id → Campagne.id
                  NPV.do_id → DonneurdOrdre.id
========================================
"""

from app import app, db, DonneurdOrdre, Campagne, NPV
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\youss\Downloads\NPV_CAMPAGNE_DO 1.xlsx"


def lire_excel(path):
    """Lit le fichier Excel et retourne une liste de tuples (campagne_nom, npv_numero, do_nom)."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))[1:]  # Skip header
    
    print(f"🔍 {len(rows)} lignes trouvées (hors en-tête)")
    print("🔍 Aperçu des 10 premières lignes :")
    for i, row in enumerate(rows[:10]):
        print(f"   Ligne {i+2}: {row}")
    
    records = []
    # ✅ Initialiser avec des valeurs vides (pas None)
    current_campagne = None
    current_do = None
    
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        
        # Colonne A : NOM_CAMPAGNE (colonne 0)
        campagne_val = row[0] if len(row) > 0 else None
        if campagne_val is not None:
            val = str(campagne_val).strip()
            if val and val.lower() != 'none':
                current_campagne = val
        
        # Colonne B : NPV (colonne 1)
        npv_val = row[1] if len(row) > 1 else None
        npv_numero = None
        if npv_val is not None:
            val = str(npv_val).strip()
            if val and val.lower() != 'none':
                npv_numero = val
        
        # Colonne C : DO (colonne 2)
        do_val = row[2] if len(row) > 2 else None
        if do_val is not None:
            val = str(do_val).strip()
            if val and val.lower() != 'none':
                current_do = val
        
        # ✅ Ne garder que les lignes qui ont un numéro NPV
        if npv_numero:
            records.append({
                'campagne': current_campagne,
                'npv': npv_numero,
                'do': current_do
            })
    
    # 🔍 DEBUG : Vérifier les premières valeurs
    print(f"\n🔍 DEBUG - Premiers enregistrements :")
    for i, r in enumerate(records[:5]):
        print(f"   {i}: campagne='{r['campagne']}' | npv='{r['npv']}' | do='{r['do']}'")
    
    return records


def importer(records):
    """Importe les données sans doublons."""
    
    stats = {
        "do_crees": 0, "do_existants": 0,
        "campagne_creees": 0, "campagne_existantes": 0,
        "npv_crees": 0, "npv_existants": 0,
        "npv_avec_do": 0, "npv_sans_do": 0,
    }
    
    # ═══════════════════════════════════════════════════
    # ÉTAPE 1 : Charger les DO existants (NE PAS EN CRÉER)
    # ═══════════════════════════════════════════════════
    print("\n📥 Chargement des Donneurs d'Ordre existants...")
    
    do_map = {}
    tous_les_dos = DonneurdOrdre.query.all()
    for do in tous_les_dos:
        do_map[do.nom] = do
        print(f"   📋 {do.nom} (ID: {do.id})")
    
    print(f"\n   Total DO en base : {len(do_map)}")
    
    # ═══════════════════════════════════════════════════
    # ÉTAPE 2 : Charger les Campagnes existantes (NE PAS EN CRÉER)
    # ═══════════════════════════════════════════════════
    print("\n📥 Chargement des Campagnes existantes...")
    
    campagne_map = {}
    toutes_les_campagnes = Campagne.query.all()
    for c in toutes_les_campagnes:
        campagne_map[c.nom] = c
        do_nom = c.do.nom if c.do else '—'
        print(f"   📋 '{c.nom}' (ID: {c.id}, DO: {do_nom})")
    
    print(f"\n   Total Campagnes en base : {len(campagne_map)}")
    
    # ═══════════════════════════════════════════════════
    # ÉTAPE 3 : Importer les NPV
    # ═══════════════════════════════════════════════════
    print("\n📥 Import des NPV...")
    
    for r in records:
        npv_numero = r['npv']
        campagne_nom = r['campagne']
        do_nom = r['do']
        
        # Vérifier doublon
        existant = NPV.query.filter_by(numero=npv_numero).first()
        if existant:
            stats["npv_existants"] += 1
            
            # ✅ Mettre à jour do_id si NULL
            if existant.do_id is None and do_nom and do_nom in do_map:
                existant.do_id = do_map[do_nom].id
                print(f"   🔄 Mise à jour do_id pour NPV existant {npv_numero} → DO {do_nom}")
            
            continue
        
        # ── Trouver campagne_id ──
        campagne_id = None
        if campagne_nom:
            if campagne_nom in campagne_map:
                campagne_id = campagne_map[campagne_nom].id
            else:
                print(f"   ⚠ Campagne '{campagne_nom}' INTROUVABLE pour NPV {npv_numero}")
        else:
            print(f"   ⚠ NPV {npv_numero} sans campagne")
        
        # ── Trouver do_id ──
        do_id = None
        if do_nom:
            if do_nom in do_map:
                do_id = do_map[do_nom].id
                stats["npv_avec_do"] += 1
            else:
                stats["npv_sans_do"] += 1
                print(f"   ⚠ DO '{do_nom}' INTROUVABLE pour NPV {npv_numero}")
        else:
            stats["npv_sans_do"] += 1
            print(f"   ⚠ NPV {npv_numero} sans DO")
        
        # ── Créer le NPV ──
        npv = NPV(
            numero=npv_numero,
            statut=None,
            campagne_id=campagne_id,
            do_id=do_id
        )
        db.session.add(npv)
        stats["npv_crees"] += 1
        
        if stats["npv_crees"] <= 10:
            print(f"   🆕 NPV {npv_numero} | Campagne: {campagne_nom or '—'} | DO: {do_nom or '—'} | do_id: {do_id}")
    
    db.session.commit()
    
    # ✅ Mise à jour des NPV existants sans do_id
    print("\n🔄 Mise à jour des NPV existants sans do_id...")
    npv_sans_do = NPV.query.filter_by(do_id=None).all()
    for npv in npv_sans_do:
        # Chercher dans les records
        for r in records:
            if r['npv'] == npv.numero and r['do'] and r['do'] in do_map:
                npv.do_id = do_map[r['do']].id
                print(f"   🔄 NPV {npv.numero} → do_id={npv.do_id} (DO: {r['do']})")
                break
    db.session.commit()
    
    return stats


def main():
    print("=" * 60)
    print("  IMPORT Excel → Base de données")
    print("  (Utilise uniquement les données existantes)")
    print("=" * 60)
    
    print("\n📖 Lecture du fichier Excel...")
    records = lire_excel(EXCEL_PATH)
    print(f"   ✅ {len(records)} enregistrements valides")
    
    print("\n📥 Insertion en base de données...")
    with app.app_context():
        stats = importer(records)
    
    print("\n" + "=" * 60)
    print("  📊 RÉSULTAT")
    print("=" * 60)
    print(f"  NPV créés      : {stats['npv_crees']}")
    print(f"  NPV existants  : {stats['npv_existants']}")
    print(f"  NPV avec do_id : {stats['npv_avec_do']}")
    if stats['npv_sans_do']:
        print(f"  ⚠ NPV sans DO trouvé : {stats['npv_sans_do']}")
    print("=" * 60)
    print("✅ Import terminé avec succès !")


if __name__ == "__main__":
    main()